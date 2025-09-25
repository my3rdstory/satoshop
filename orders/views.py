from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.conf import settings
import csv
import json
import logging

from stores.models import Store
from products.models import Product, ProductOption, ProductOptionChoice
from .models import Cart, CartItem, Order, OrderItem, PurchaseHistory, Invoice
from .services import CartService
from stores.decorators import store_owner_required
from ln_payment.blink_service import get_blink_service_for_store

logger = logging.getLogger(__name__)


def _format_text_for_csv(value):
    """CSV 다운로드 시 Excel에서 일반 텍스트로 인식되도록 포맷팅"""
    if value in (None, ''):
        return ''

    text_value = str(value)
    # Excel이 수식으로 처리하지 않도록 큰따옴표 이스케이프 후 텍스트 서식으로 감싼다.
    escaped = text_value.replace('"', '""')
    return f'="{escaped}"'


def calculate_store_totals(store_obj, store_items):
    """스토어별 소계, 배송비, 무조건 무료 배송 여부 계산"""
    subtotal = 0
    apply_override = store_obj.shipping_fee_mode == 'flat'
    has_items = False

    for item in store_items:
        has_items = True
        if isinstance(item, dict):
            subtotal += item.get('total_price', 0) or 0
            item_force_free = item.get('force_free_shipping', False)
        else:
            subtotal += getattr(item, 'total_price', 0) or 0
            product = getattr(item, 'product', None)
            item_force_free = getattr(product, 'force_free_shipping', False)

        apply_override = apply_override and item_force_free

    if not has_items:
        apply_override = False

    shipping_fee = 0 if apply_override else store_obj.get_shipping_fee_sats(subtotal)
    return subtotal, shipping_fee, apply_override


@login_required
def cart_view(request):
    """장바구니 보기"""
    cart_service = CartService(request)
    cart_items = cart_service.get_cart_items()
    cart_summary = cart_service.get_cart_summary()
    
    # store_base.html을 위한 기본 store 설정
    # 장바구니에 상품이 있으면 첫 번째 상품의 스토어를 사용
    # 없으면 기본값 설정
    store = None
    if cart_items:
        # 첫 번째 상품의 스토어 정보 가져오기
        first_item = cart_items[0]
        try:
            store = Store.objects.filter(store_id=first_item['store_id'], deleted_at__isnull=True).first()
        except Exception:
            store = None
    
    if not store:
        # 기본 스토어 정보 (빈 장바구니일 때)
        class DummyStore:
            store_id = ''
            store_name = 'SatoShop'
            store_description = ''
            owner_name = ''
            owner_phone = ''
            owner_email = ''
            chat_channel = ''
        store = DummyStore()
    
    # 기존 Cart 객체와 호환되는 더미 객체 생성
    class DummyCart:
        def __init__(self, total_amount, total_items):
            self.total_amount = total_amount
            self.total_items = total_items
    
    dummy_cart = DummyCart(cart_summary['total_amount'], cart_summary['total_items'])
    
    # 원화 연동 상품 여부 확인
    has_krw_products = False
    current_exchange_rate = None
    
    if cart_items:
        # 장바구니에 원화 연동 상품이 있는지 확인
        for item in cart_items:
            try:
                product = Product.objects.get(id=item['product_id'])
                if product.price_krw:
                    has_krw_products = True
                    break
            except Product.DoesNotExist:
                continue
        
        # 원화 연동 상품이 있으면 현재 환율 정보 가져오기
        if has_krw_products:
            try:
                from myshop.models import ExchangeRate
                exchange_rate = ExchangeRate.objects.latest('created_at')
                current_exchange_rate = exchange_rate.btc_krw_rate
            except ExchangeRate.DoesNotExist:
                pass
    
    context = {
        'cart_items': cart_items,
        'cart': dummy_cart,
        'store': store,
        'cart_items_count': cart_summary['total_items'],
        'cart_total_amount': cart_summary['total_amount'],
        'has_krw_products': has_krw_products,
        'current_exchange_rate': current_exchange_rate,
    }
    return render(request, 'orders/cart.html', context)


def cart_api(request):
    """장바구니 내용을 JSON으로 반환하는 API"""
    if not request.user.is_authenticated:
        return JsonResponse({
            'success': False,
            'error': 'login_required',
            'message': '로그인이 필요합니다.'
        }, status=401)

    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'GET 요청만 허용됩니다.'})
    
    try:
        cart_service = CartService(request)
        cart_items = cart_service.get_cart_items()
        cart_summary = cart_service.get_cart_summary()
        
        # 스토어별로 그룹화
        stores_data = {}
        for item in cart_items:
            store_id = item['store_id']
            if store_id not in stores_data:
                stores_data[store_id] = {
                    'store_name': item['store_name'],
                    'store_id': store_id,
                    'items': []
                }
            
            stores_data[store_id]['items'].append({
                'id': item['id'],
                'product': {
                    'title': item['product_title'],
                    'images': {'first': {'file_url': item['product_image_url']} if item['product_image_url'] else None}
                },
                'quantity': item['quantity'],
                'total_price': item['total_price'],
                'options_display': item['options_display']
            })
        
        cart_items_grouped = list(stores_data.values())
        
        return JsonResponse({
            'success': True,
            'cart_items': cart_items_grouped,
            'total_items': cart_summary['total_items'],
            'total_amount': cart_summary['total_amount']
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def check_cart_store_conflict(request):
    """장바구니 스토어 충돌 여부 확인"""
    if not request.user.is_authenticated:
        return JsonResponse({
            'success': False,
            'error': 'login_required',
            'message': '로그인이 필요합니다.'
        }, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST 요청만 허용됩니다.'})
    
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        
        if not product_id:
            return JsonResponse({'success': False, 'error': '상품 ID가 필요합니다.'})
        
        try:
            from products.models import Product
            product = Product.objects.get(id=product_id, is_active=True)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': '상품을 찾을 수 없습니다.'})
        
        cart_service = CartService(request)
        existing_items = cart_service.get_cart_items()
        
        if existing_items:
            # 기존 장바구니에 있는 스토어들 확인
            existing_stores = set(item['store_id'] for item in existing_items)
            current_store_id = product.store.store_id
            
            # 다른 스토어의 상품이 이미 있는 경우
            if current_store_id not in existing_stores:
                existing_store_names = set(item['store_name'] for item in existing_items)
                return JsonResponse({
                    'success': True,
                    'has_conflict': True,
                    'current_store': product.store.store_name,
                    'existing_stores': list(existing_store_names),
                    'message': f'장바구니에 다른 스토어({", ".join(existing_store_names)})의 상품이 있습니다.'
                })
        
        return JsonResponse({
            'success': True,
            'has_conflict': False
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def add_to_cart(request):
    """장바구니에 상품 추가"""
    if not request.user.is_authenticated:
        return JsonResponse({
            'success': False,
            'error': 'login_required',
            'message': '로그인이 필요합니다.'
        }, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST 요청만 허용됩니다.'})
    
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        quantity = int(data.get('quantity', 1))
        selected_options = data.get('selected_options', {})
        force_replace = data.get('force_replace', False)  # 장바구니 교체 강제 여부
        
        cart_service = CartService(request)
        result = cart_service.add_to_cart(product_id, quantity, selected_options, force_replace)
        
        if result['success']:
            cart_summary = cart_service.get_cart_summary()
            return JsonResponse({
                'success': True,
                'message': result['message'],
                'cart_count': cart_summary['total_items'],
                'cart_total_amount': cart_summary['total_amount']
            })
        else:
            return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def remove_from_cart(request, item_id=None):
    """장바구니에서 상품 제거"""
    if not request.user.is_authenticated:
        if request.method == 'POST':
            return JsonResponse({
                'success': False,
                'error': 'login_required',
                'message': '로그인이 필요합니다.'
            }, status=401)
        login_url = f"{settings.LOGIN_URL}?next={request.get_full_path()}"
        return redirect(login_url)

    if request.method == 'POST':
        try:
            # JSON 요청 처리
            if request.content_type == 'application/json':
                import json
                data = json.loads(request.body)
                item_id = data.get('item_id')
            
            if not item_id:
                return JsonResponse({'success': False, 'error': '아이템 ID가 필요합니다.'})
            
            cart_service = CartService(request)
            result = cart_service.remove_from_cart(item_id)
            return JsonResponse(result)
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET 요청 (기존 방식 유지)
    if item_id:
        cart_service = CartService(request)
        result = cart_service.remove_from_cart(item_id)
        if result['success']:
            messages.success(request, result['message'])
        else:
            messages.error(request, result['error'])
    
    return redirect('orders:cart_view')


def update_cart_item(request, item_id):
    """장바구니 상품 수량 업데이트"""
    if not request.user.is_authenticated:
        return JsonResponse({
            'success': False,
            'error': 'login_required',
            'message': '로그인이 필요합니다.'
        }, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST 요청만 허용됩니다.'})
    
    try:
        quantity = int(request.POST.get('quantity', 1))
        
        cart_service = CartService(request)
        result = cart_service.update_cart_item(item_id, quantity)
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@store_owner_required
def order_management(request, store_id):
    """주문 관리"""
    store = get_object_or_404(Store, store_id=store_id, owner=request.user, deleted_at__isnull=True)
    
    # 현재 월 파라미터 처리
    from datetime import datetime
    import calendar
    
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    
    # 선택된 월의 시작과 끝
    month_start = timezone.datetime(year, month, 1, tzinfo=timezone.get_current_timezone())
    if month == 12:
        month_end = timezone.datetime(year + 1, 1, 1, tzinfo=timezone.get_current_timezone())
    else:
        month_end = timezone.datetime(year, month + 1, 1, tzinfo=timezone.get_current_timezone())
    
    # 스토어 생성월 확인
    store_created = store.created_at
    store_year = store_created.year
    store_month = store_created.month
    
    # 현재 월 확인
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    
    # 이전/다음 월 계산
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    # 이전/다음 버튼 표시 여부
    show_prev = not (prev_year < store_year or (prev_year == store_year and prev_month < store_month))
    show_next = not (next_year > current_year or (next_year == current_year and next_month > current_month))
    
    # 선택된 월 통계
    monthly_orders = Order.objects.filter(
        items__product__store=store,
        created_at__gte=month_start,
        created_at__lt=month_end
    ).distinct()
    
    monthly_orders_count = monthly_orders.count()
    monthly_revenue = monthly_orders.aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    
    # 상품별 판매 현황 (선택된 월 기준)
    from django.db.models import F
    products_with_orders = Product.objects.filter(
        store=store
    ).annotate(
        total_orders=Count('orderitem__order', distinct=True, filter=Q(
            orderitem__order__created_at__gte=month_start,
            orderitem__order__created_at__lt=month_end
        )),
        total_quantity=Sum('orderitem__quantity', filter=Q(
            orderitem__order__created_at__gte=month_start,
            orderitem__order__created_at__lt=month_end
        )),
        total_revenue=Sum(
            F('orderitem__quantity') * (F('orderitem__product_price') + F('orderitem__options_price')),
            filter=Q(
                orderitem__order__created_at__gte=month_start,
                orderitem__order__created_at__lt=month_end
            )
        )
    ).order_by('-total_revenue', '-created_at')
    
    context = {
        'store': store,
        'monthly_orders_count': monthly_orders_count,
        'monthly_revenue': monthly_revenue,
        'products_with_orders': products_with_orders,
        'current_year': year,
        'current_month': month,
        'current_month_name': f'{month}월',
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'show_prev': show_prev,
        'show_next': show_next,
    }
    return render(request, 'orders/order_management.html', context)


@login_required
@store_owner_required
def product_orders(request, store_id, product_id):
    """특정 상품의 주문 목록"""
    import calendar
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta
    
    store = get_object_or_404(Store, store_id=store_id, owner=request.user, deleted_at__isnull=True)
    product = get_object_or_404(Product, id=product_id, store=store)
    
    # 기본 쿼리셋
    order_items = OrderItem.objects.filter(
        product=product
    ).select_related('order')
    
    # 현재 날짜 정보
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    
    # 이번달 범위
    current_month_start = datetime(current_year, current_month, 1)
    if current_month == 12:
        current_month_end = datetime(current_year + 1, 1, 1)
    else:
        current_month_end = datetime(current_year, current_month + 1, 1)
    
    # 지난달 범위
    last_month_date = now - relativedelta(months=1)
    last_month_year = last_month_date.year
    last_month_month = last_month_date.month
    last_month_start = datetime(last_month_year, last_month_month, 1)
    if last_month_month == 12:
        last_month_end = datetime(last_month_year + 1, 1, 1)
    else:
        last_month_end = datetime(last_month_year, last_month_month + 1, 1)
    
    # 필터 파라미터 처리 (기본값을 현재 월로 설정)
    filter_type = request.GET.get('filter', 'this_month')  # all, this_month, last_month, custom
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # 날짜 필터링 적용
    if filter_type == 'this_month':
        order_items = order_items.filter(
            order__created_at__gte=current_month_start,
            order__created_at__lt=current_month_end
        )
    elif filter_type == 'last_month':
        order_items = order_items.filter(
            order__created_at__gte=last_month_start,
            order__created_at__lt=last_month_end
        )
    elif filter_type == 'custom' and start_date and end_date:
        try:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)  # 종료일 포함
            order_items = order_items.filter(
                order__created_at__gte=start_datetime,
                order__created_at__lt=end_datetime
            )
        except ValueError:
            # 날짜 형식이 잘못된 경우 전체 조회
            pass
    
    # 정렬 적용 (최신순)
    order_items = order_items.order_by('-order__created_at')
    
    # 페이지네이션 (전체 필터일 때만 적용)
    page_obj = None
    page_numbers = []
    if filter_type == 'all':
        paginator = Paginator(order_items, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # 페이지네이션을 위한 페이지 번호 리스트 계산 (항상 5개)
        def get_page_numbers(current_page, total_pages):
            """현재 페이지를 중심으로 5개의 페이지 번호 리스트 반환"""
            if total_pages <= 5:
                return list(range(1, total_pages + 1))
            
            # 현재 페이지를 중심으로 ±2 범위
            start = max(1, current_page - 2)
            end = min(total_pages, current_page + 2)
            
            # 시작이나 끝에 치우쳐있을 때 조정
            if end - start < 4:  # 5개가 되도록 조정
                if start == 1:
                    end = min(total_pages, start + 4)
                elif end == total_pages:
                    start = max(1, end - 4)
            
            return list(range(start, end + 1))
        
        page_numbers = get_page_numbers(page_obj.number, page_obj.paginator.num_pages) if page_obj.has_other_pages() else []
    
    # 필터링된 결과의 통계 계산 (페이지네이션이 적용된 경우와 아닌 경우 모두 전체 데이터 기준)
    if filter_type == 'all' and page_obj:
        # 전체 필터일 때는 전체 데이터 기준으로 통계 계산
        filtered_stats = order_items.aggregate(
            total_orders=Count('order', distinct=True),
            total_quantity=Sum('quantity'),
            total_revenue=Sum(F('quantity') * (F('product_price') + F('options_price')))
        )
    else:
        # 다른 필터일 때는 필터링된 전체 데이터 기준으로 통계 계산
        filtered_stats = order_items.aggregate(
            total_orders=Count('order', distinct=True),
            total_quantity=Sum('quantity'),
            total_revenue=Sum(F('quantity') * (F('product_price') + F('options_price')))
        )
    
    context = {
        'store': store,
        'product': product,
        'page_obj': page_obj,
        'page_numbers': page_numbers,
        'order_items': order_items if filter_type != 'all' else None,  # 페이지네이션이 없을 때 전체 데이터
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'current_month_name': f'{current_month}월',
        'current_year': current_year,
        'last_month_name': f'{last_month_month}월',
        'last_month_year': last_month_year,
        'filtered_stats': filtered_stats,
    }
    return render(request, 'orders/product_orders.html', context)


@login_required
@store_owner_required
def export_orders_csv(request, store_id):
    """주문 데이터 엑셀 내보내기 (배송 정보 포함)"""
    store = get_object_or_404(Store, store_id=store_id, owner=request.user, deleted_at__isnull=True)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "주문 목록"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        headers = [
            '주문번호', '상품명', '수량', '단가', '총액', '주문일시', '상태',
            '주문자명', '연락처', '이메일', '우편번호', '주소', '상세주소', '요청사항'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 조회
        order_items = OrderItem.objects.filter(
            product__store=store
        ).select_related('order', 'product').order_by('-order__created_at')
        
        # 데이터 작성
        for row, item in enumerate(order_items, 2):
            order = item.order
            ws.cell(row=row, column=1, value=order.order_number)
            ws.cell(row=row, column=2, value=item.product_title)
            ws.cell(row=row, column=3, value=item.quantity)
            ws.cell(row=row, column=4, value=item.product_price)
            ws.cell(row=row, column=5, value=item.total_price)
            ws.cell(row=row, column=6, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=7, value=order.get_status_display())
            ws.cell(row=row, column=8, value=order.buyer_name)
            phone_cell = ws.cell(row=row, column=9, value=str(order.buyer_phone or ''))
            phone_cell.number_format = '@'
            ws.cell(row=row, column=10, value=order.buyer_email)
            ws.cell(row=row, column=11, value=order.shipping_postal_code)
            ws.cell(row=row, column=12, value=order.shipping_address)
            ws.cell(row=row, column=13, value=order.shipping_detail_address)
            ws.cell(row=row, column=14, value=order.order_memo)
        
        # 열 너비 자동 조정
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # 메모리에 엑셀 파일 저장
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # HTTP 응답 생성
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{store.store_id}_orders_with_shipping.xlsx"'
        
        return response
    
    except ImportError:
        # openpyxl이 없으면 기존 CSV 방식 사용
        logger.warning("openpyxl 라이브러리가 없어 CSV로 다운로드됩니다.")
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{store.store_id}_orders.csv"'
        response.write('\ufeff')  # UTF-8 BOM for Excel
        
        writer = csv.writer(response)
        writer.writerow([
            '주문번호', '상품명', '수량', '단가', '총액', '주문일시', '상태',
            '주문자명', '연락처', '이메일', '우편번호', '주소', '상세주소', '요청사항'
        ])
        
        order_items = OrderItem.objects.filter(
            product__store=store
        ).select_related('order', 'product').order_by('-order__created_at')
        
        for item in order_items:
            order = item.order
            writer.writerow([
                order.order_number,
                item.product_title,
                item.quantity,
                item.product_price,
                item.total_price,
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                order.get_status_display(),
                order.buyer_name,
                _format_text_for_csv(order.buyer_phone),
                order.buyer_email,
                order.shipping_postal_code,
                order.shipping_address,
                order.shipping_detail_address,
                order.order_memo,
            ])
        
        return response


@login_required
@store_owner_required
def export_product_orders_csv(request, store_id, product_id):
    """특정 상품의 주문 데이터 CSV 다운로드 (현재 화면 필터 반영)"""
    from .products_orders_csv_download import export_product_orders_csv as export_csv
    
    store = get_object_or_404(Store, store_id=store_id, owner=request.user, deleted_at__isnull=True)
    product = get_object_or_404(Product, id=product_id, store=store)
    
    # 새로운 모듈의 함수를 사용하여 CSV 다운로드
    return export_csv(request, store, product)


@login_required
def my_purchases(request):
    """구매 내역 목록"""
    purchases = PurchaseHistory.objects.filter(user=request.user).select_related(
        'order', 'order__store'
    ).prefetch_related(
        'order__items', 'order__items__product', 'order__items__product__images'
    ).order_by('-purchase_date')
    
    context = {
        'purchases': purchases,
    }
    return render(request, 'orders/my_purchases.html', context)


@login_required
def purchase_detail(request, order_number):
    """구매 상세 정보"""
    order = get_object_or_404(Order, order_number=order_number, user=request.user)
    
    context = {
        'order': order,
    }
    return render(request, 'orders/purchase_detail.html', context)


@login_required
def shipping_info(request):
    """배송 정보 입력 페이지"""
    cart_service = CartService(request)
    cart_items = cart_service.get_cart_items()
    
    if not cart_items:
        messages.error(request, '장바구니가 비어있습니다.')
        return redirect('myshop:home')
    
    if request.method == 'POST':
        # 배송 정보를 세션에 저장
        shipping_data = {
            'buyer_name': request.POST.get('buyer_name', '').strip(),
            'buyer_phone': request.POST.get('buyer_phone', '').strip(),
            'buyer_email': request.POST.get('buyer_email', '').strip(),
            'shipping_postal_code': request.POST.get('shipping_postal_code', '').strip(),
            'shipping_address': request.POST.get('shipping_address', '').strip(),
            'shipping_detail_address': request.POST.get('shipping_detail_address', '').strip(),
            'order_memo': request.POST.get('order_memo', '').strip(),
        }
        
        # 필수 필드 검증
        required_fields = ['buyer_name', 'buyer_phone', 'buyer_email', 'shipping_postal_code', 'shipping_address']
        missing_fields = []
        for field in required_fields:
            if not shipping_data[field]:
                missing_fields.append(field)
        
        if missing_fields:
            messages.error(request, '필수 정보를 모두 입력해주세요.')
            
            # 스토어별로 그룹화하여 배송비 계산 (에러 시에도 필요)
            from itertools import groupby
            stores_with_items = []
            total_shipping_fee = 0
            
            cart_items_sorted = cart_items.order_by('product__store__store_name', '-added_at')
            
            for store, items in groupby(cart_items_sorted, key=lambda x: x.product.store):
                store_items = list(items)
                store_subtotal, store_shipping_fee, override_free = calculate_store_totals(store, store_items)
                store_total = store_subtotal + store_shipping_fee
                total_shipping_fee += store_shipping_fee

                stores_with_items.append({
                    'store': store,
                    'items': store_items,
                    'subtotal': store_subtotal,
                    'shipping_fee': store_shipping_fee,
                    'total': store_total,
                    'shipping_info': store.get_shipping_fee_display(),
                    'force_free_override': override_free,
                })
            
            # 전체 총액 계산
            subtotal_amount = sum(store_data['subtotal'] for store_data in stores_with_items)
            total_amount = subtotal_amount + total_shipping_fee
            
            context = {
                'cart': cart,
                'cart_items': cart_items,
                'stores_with_items': stores_with_items,
                'subtotal_amount': subtotal_amount,
                'total_shipping_fee': total_shipping_fee,
                'total_amount': total_amount,
                'form_data': shipping_data,
                'store': cart_items[0].product.store if cart_items else None,
            }
            return render(request, 'orders/shipping_info.html', context)
        
        # 세션에 배송 정보 저장
        request.session['shipping_data'] = shipping_data
        
        # 총 금액이 0인 경우 바로 주문 생성 후 완료 화면으로 이동
        cart_service = CartService(request)
        cart_items = cart_service.get_cart_items()
        
        # 총 금액 계산
        stores_data = {}
        total_shipping_fee = 0
        
        for item in cart_items:
            store_id = item['store_id']
            if store_id not in stores_data:
                stores_data[store_id] = {
                    'store_name': item['store_name'],
                    'store_id': store_id,
                    'items': []
                }
            stores_data[store_id]['items'].append(item)
        
        for store_id, store_data in stores_data.items():
            store_items = store_data['items']
            store_obj = Store.objects.filter(store_id=store_id, deleted_at__isnull=True).first()
            if not store_obj:
                continue

            _, store_shipping_fee, _ = calculate_store_totals(store_obj, store_items)
            total_shipping_fee += store_shipping_fee
        
        # 전체 총액 계산
        subtotal_amount = sum(item['total_price'] for item in cart_items)
        total_amount = subtotal_amount + total_shipping_fee
        
        if total_amount == 0:
            # 무료 주문인 경우 바로 주문 생성
            try:
                import uuid
                from django.utils import timezone
                
                # 임시 결제 해시 생성 (무료 주문용)
                payment_hash = f"FREE-{timezone.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:8].upper()}"
                
                # 주문 생성
                order_result = create_order_from_cart_service(request, payment_hash, shipping_data)
                
                # 세션에서 배송 정보 삭제
                if 'shipping_data' in request.session:
                    del request.session['shipping_data']
                
                # 주문 완료 페이지로 이동
                messages.success(request, '무료 주문이 완료되었습니다!')
                return redirect('orders:checkout_complete', order_number=order_result['primary_order_number'])
                
            except Exception as e:
                messages.error(request, f'주문 처리 중 오류가 발생했습니다: {str(e)}')
                return redirect('orders:shipping_info')
        
        return redirect('orders:checkout')
    
    # 스토어별로 그룹화하여 배송비 계산
    stores_with_items = []
    total_shipping_fee = 0
    
    # 스토어별로 그룹화
    stores_data = {}
    for item in cart_items:
        store_id = item['store_id']
        if store_id not in stores_data:
            stores_data[store_id] = {
                'store_name': item['store_name'],
                'store_id': store_id,
                'items': []
            }
        stores_data[store_id]['items'].append(item)
    
    for store_id, store_data in stores_data.items():
        store_items = store_data['items']
        store_obj = Store.objects.filter(store_id=store_id, deleted_at__isnull=True).first()
        if not store_obj:
            continue

        store_subtotal, store_shipping_fee, override_free = calculate_store_totals(store_obj, store_items)
        store_total = store_subtotal + store_shipping_fee
        total_shipping_fee += store_shipping_fee

        stores_with_items.append({
            'store': store_obj,
            'items': store_items,
            'subtotal': store_subtotal,
            'shipping_fee': store_shipping_fee,
            'total': store_total,
            'shipping_info': store_obj.get_shipping_fee_display(),
            'force_free_override': override_free,
        })
    
    # 전체 총액 계산
    subtotal_amount = sum(store_data['subtotal'] for store_data in stores_with_items)
    total_amount = subtotal_amount + total_shipping_fee
    
    # store_base.html을 위한 기본 store 설정
    store = stores_with_items[0]['store'] if stores_with_items else None
    
    # 원화 연동 상품 여부 확인
    has_krw_products = False
    if cart_items:
        # 장바구니에 원화 연동 상품이 있는지 확인
        for item in cart_items:
            try:
                product = Product.objects.get(id=item['product_id'])
                if product.price_krw:
                    has_krw_products = True
                    break
            except Product.DoesNotExist:
                continue
    
    # 기존 Cart 객체와 호환되는 더미 객체 생성
    cart_summary = cart_service.get_cart_summary()
    class DummyCart:
        def __init__(self, total_amount, total_items):
            self.total_amount = total_amount
            self.total_items = total_items
    
    dummy_cart = DummyCart(cart_summary['total_amount'], cart_summary['total_items'])
    
    context = {
        'cart': dummy_cart,
        'cart_items': cart_items,
        'stores_with_items': stores_with_items,
        'subtotal_amount': subtotal_amount,
        'total_shipping_fee': total_shipping_fee,
        'total_amount': total_amount,
        'store': store,
        'has_krw_products': has_krw_products,
    }
    return render(request, 'orders/shipping_info.html', context)


@login_required
def checkout(request):
    """주문/결제 페이지"""
    # 배송 정보가 세션에 있는지 확인
    if 'shipping_data' not in request.session:
        messages.error(request, '배송 정보를 먼저 입력해주세요.')
        return redirect('orders:shipping_info')
    
    cart_service = CartService(request)
    cart_items = cart_service.get_cart_items()
    
    if not cart_items:
        messages.error(request, '장바구니가 비어있습니다.')
        return redirect('myshop:home')
    
    # 스토어별로 그룹화
    stores_with_items = []
    total_shipping_fee = 0
    
    if settings.DEBUG:
        logger.debug(f"[CHECKOUT] 장바구니 아이템 수: {len(cart_items)}")
    
    # 스토어별로 그룹화
    stores_data = {}
    for item in cart_items:
        store_id = item['store_id']
        if store_id not in stores_data:
            stores_data[store_id] = {
                'store_name': item['store_name'],
                'store_id': store_id,
                'items': []
            }
        stores_data[store_id]['items'].append(item)
    
    for store_id, store_data in stores_data.items():
        store_items = store_data['items']
        store_obj = Store.objects.filter(store_id=store_id, deleted_at__isnull=True).first()
        if not store_obj:
            continue

        store_subtotal, store_shipping_fee, override_free = calculate_store_totals(store_obj, store_items)
        store_total = store_subtotal + store_shipping_fee
        total_shipping_fee += store_shipping_fee

        if settings.DEBUG:
            logger.debug(f"[CHECKOUT] 스토어: {store_obj.store_name}")
            logger.debug(f"[CHECKOUT]   - 상품 수: {len(store_items)}")
            logger.debug(f"[CHECKOUT]   - 상품 소계: {store_subtotal} sats")
            logger.debug(f"[CHECKOUT]   - 배송비: {store_shipping_fee} sats (스토어 정책)")
            logger.debug(f"[CHECKOUT]   - 스토어 총액: {store_total} sats")
            for item in store_items:
                logger.debug(f"[CHECKOUT]     * {item['product_title']}: {item['quantity']}개 x {item['unit_price']} = {item['total_price']} sats")
        
        stores_with_items.append({
            'store': store_obj,
            'items': store_items,
            'subtotal': store_subtotal,
            'shipping_fee': store_shipping_fee,
            'total': store_total,
            'shipping_info': store_obj.get_shipping_fee_display(),
            'force_free_override': override_free,
        })
    
    # 전체 총액 계산
    subtotal_amount = sum(store_data['subtotal'] for store_data in stores_with_items)
    total_amount = subtotal_amount + total_shipping_fee
    
    if settings.DEBUG:
        logger.debug(f"[CHECKOUT] 전체 계산 결과:")
        logger.debug(f"[CHECKOUT]   - 상품 총액: {subtotal_amount} sats")
        logger.debug(f"[CHECKOUT]   - 배송비 총액: {total_shipping_fee} sats")
        logger.debug(f"[CHECKOUT]   - 최종 총액: {total_amount} sats")
    
    # store_base.html을 위한 기본 store 설정
    store = stores_with_items[0]['store'] if stores_with_items else None
    
    # 원화 연동 상품 여부 확인
    has_krw_products = False
    if cart_items:
        # 장바구니에 원화 연동 상품이 있는지 확인
        for item in cart_items:
            try:
                product = Product.objects.get(id=item['product_id'])
                if product.price_krw:
                    has_krw_products = True
                    break
            except Product.DoesNotExist:
                continue
    
    # 세션에서 배송 정보 가져오기
    shipping_data = request.session.get('shipping_data', {})
    
    # 기존 Cart 객체와 호환되는 더미 객체 생성
    cart_summary = cart_service.get_cart_summary()
    class DummyCart:
        def __init__(self, total_amount, total_items):
            self.total_amount = total_amount
            self.total_items = total_items
    
    dummy_cart = DummyCart(cart_summary['total_amount'], cart_summary['total_items'])
    
    context = {
        'cart': dummy_cart,
        'cart_items': cart_items,
        'stores_with_items': stores_with_items,
        'subtotal_amount': subtotal_amount,
        'total_shipping_fee': total_shipping_fee,
        'total_amount': total_amount,
        'shipping_data': shipping_data,
        'store': store,
        'has_krw_products': has_krw_products,
    }
    return render(request, 'orders/checkout.html', context)


@login_required
def checkout_complete(request, order_number):
    """주문 완료 페이지"""
    try:
        # 먼저 주문 번호로만 조회
        primary_order = get_object_or_404(Order, order_number=order_number)
        
        # 로그인된 사용자인 경우 소유권 확인
        if request.user.is_authenticated:
            # 현재 로그인한 사용자의 주문인지 확인
            if primary_order.user != request.user:
                # 다른 사용자의 주문이라면 접근 거부
                raise Http404("주문을 찾을 수 없습니다.")
            
            # 같은 결제 해시로 생성된 모든 주문 가져오기
            # payment_id가 비어있거나 None인 경우 단일 주문만 처리
            if primary_order.payment_id:
                all_orders = Order.objects.filter(
                    user=request.user,
                    payment_id=primary_order.payment_id
                ).prefetch_related('items__product__images')
            else:
                # 무료 상품 등으로 payment_id가 없는 경우 단일 주문만 처리
                all_orders = Order.objects.filter(
                    id=primary_order.id
                ).prefetch_related('items__product__images')
        else:
            # 비로그인 사용자의 경우
            # 익명 사용자 주문인지 확인
            if primary_order.user.username != 'anonymous_guest':
                # 일반 사용자의 주문이라면 접근 거부
                raise Http404("주문을 찾을 수 없습니다.")
            
            # 같은 결제 해시로 생성된 모든 주문 가져오기
            # payment_id가 비어있거나 None인 경우 단일 주문만 처리
            if primary_order.payment_id:
                all_orders = Order.objects.filter(
                    payment_id=primary_order.payment_id
                ).prefetch_related('items__product__images')
            else:
                # 무료 상품 등으로 payment_id가 없는 경우 단일 주문만 처리
                all_orders = Order.objects.filter(
                    id=primary_order.id
                ).prefetch_related('items__product__images')
    except Order.DoesNotExist:
        # 주문이 존재하지 않는 경우
        raise Http404("주문을 찾을 수 없습니다.")
    
    # 주문별 무조건 무료 배송 여부 계산 (스토어 유료 배송 + 모든 상품 강제 무료)
    for order in all_orders:
        if order.store.shipping_fee_mode == 'flat' and order.shipping_fee == 0:
            order.force_free_override = all(
                getattr(item.product, 'force_free_shipping', False)
                for item in order.items.all()
            )
        else:
            order.force_free_override = False

    # 전체 통계 계산
    total_amount = sum(order.total_amount for order in all_orders)
    total_subtotal = sum(order.subtotal for order in all_orders)
    total_shipping_fee = sum(order.shipping_fee for order in all_orders)
    total_items = sum(order.items.count() for order in all_orders)
    
    # 모든 주문 아이템을 하나의 리스트로 합치기 (주문 정보도 함께 저장)
    all_order_items = []
    for order in all_orders:
        for item in order.items.all():
            # 아이템에 주문 정보를 추가
            item.order = order
            all_order_items.append(item)
    
    context = {
        'primary_order': primary_order,
        'all_orders': all_orders,
        'all_order_items': all_order_items,
        'total_amount': total_amount,
        'total_subtotal': total_subtotal,
        'total_shipping_fee': total_shipping_fee,
        'total_items': total_items,
        'total_orders': len(all_orders),
        'store': primary_order.store,  # 기본 스토어 (store_base.html용)
    }
    return render(request, 'orders/checkout_complete.html', context)


@require_POST
def create_checkout_invoice(request):
    """결제용 인보이스 생성"""
    try:
        if settings.DEBUG:
            logger.debug("create_checkout_invoice 호출")
        
        # 요청 본문 파싱
        try:
            data = json.loads(request.body)
            if settings.DEBUG:
                logger.debug(f"인보이스 생성 요청 데이터: {data}")
        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'error': f'JSON 파싱 오류: {str(e)}'}, status=400)
        
        # 장바구니 확인
        cart_service = CartService(request)
        cart_items = cart_service.get_cart_items()
        
        if not cart_items:
            return JsonResponse({'success': False, 'error': '장바구니가 비어있습니다.'}, status=400)
        
        # 스토어별로 그룹화하여 총액 계산
        stores_data = {}
        total_shipping_fee = 0
        
        if settings.DEBUG:
            logger.debug(f"[INVOICE] 장바구니 아이템 수: {len(cart_items)}")
        
        # 스토어별로 그룹화
        for item in cart_items:
            store_id = item['store_id']
            if store_id not in stores_data:
                stores_data[store_id] = {
                    'store_name': item['store_name'],
                    'store_id': store_id,
                    'items': []
                }
            stores_data[store_id]['items'].append(item)
        
        stores_with_items = []
        for store_id, store_data in stores_data.items():
            store_items = store_data['items']
            store_obj = Store.objects.filter(store_id=store_id, deleted_at__isnull=True).first()
            if not store_obj:
                continue

            store_subtotal, store_shipping_fee, override_free = calculate_store_totals(store_obj, store_items)
            store_total = store_subtotal + store_shipping_fee
            total_shipping_fee += store_shipping_fee
            
            if settings.DEBUG:
                logger.debug(f"[INVOICE] 스토어: {store_obj.store_name}")
                logger.debug(f"[INVOICE]   - 상품 수: {len(store_items)}")
                logger.debug(f"[INVOICE]   - 상품 소계: {store_subtotal} sats")
                logger.debug(f"[INVOICE]   - 배송비: {store_shipping_fee} sats (스토어 정책)")
                logger.debug(f"[INVOICE]   - 스토어 총액: {store_total} sats")
                for item in store_items:
                    logger.debug(f"[INVOICE]     * {item['product_title']}: {item['quantity']}개 x {item['unit_price']} = {item['total_price']} sats")
            
            stores_with_items.append({
                'store': store_obj,
                'items': store_items,
                'subtotal': store_subtotal,
                'shipping_fee': store_shipping_fee,
                'total': store_total,
                'shipping_info': store_obj.get_shipping_fee_display(),
                'force_free_override': override_free,
            })
        
        # 전체 총액 계산
        subtotal_amount = sum(store_data['subtotal'] for store_data in stores_with_items)
        total_amount = subtotal_amount + total_shipping_fee
        
        if settings.DEBUG:
            logger.debug(f"[INVOICE] 전체 계산 결과:")
            logger.debug(f"[INVOICE]   - 상품 총액: {subtotal_amount} sats")
            logger.debug(f"[INVOICE]   - 배송비 총액: {total_shipping_fee} sats")
            logger.debug(f"[INVOICE]   - 최종 총액: {total_amount} sats")
        
        # 첫 번째 스토어의 블링크 API 사용 (단일 스토어 제약으로 항상 하나의 스토어만 존재)
        first_store = stores_with_items[0]['store']
        
        # 메모 생성 - 제품 정보 포함
        memo_parts = []
        
        # 각 상품에 대한 정보를 수집
        for store_data in stores_with_items:
            for item in store_data['items']:
                product_name = item['product_title']
                quantity = item['quantity']
                price = item['total_price']
                memo_parts.append(f"제품: {product_name} / 수량: {quantity}개 / 금액: {price} sats")
        
        # 메모 조합 (최대 3개 상품까지만 표시하고 나머지는 생략)
        if len(memo_parts) <= 3:
            memo = " | ".join(memo_parts)
        else:
            memo = " | ".join(memo_parts[:3]) + f" | 외 {len(memo_parts) - 3}개 상품"
        
        # 총액 정보 추가
        memo += f" | 총 {total_amount} sats"
        
        # BlinkAPIService 초기화
        try:
            blink_service = get_blink_service_for_store(first_store)
            if settings.DEBUG:
                logger.debug("BlinkAPIService 초기화 성공")
        except Exception as e:
            if settings.DEBUG:
                logger.error(f"BlinkAPIService 초기화 실패: {str(e)}")
            return JsonResponse({'success': False, 'error': f'Blink API 서비스 초기화 실패: {str(e)}'}, status=500)
        
        # 인보이스 생성
        if settings.DEBUG:
            logger.debug(f"[INVOICE] 인보이스 생성 시작: amount={total_amount} sats, memo={memo}")
        
        result = blink_service.create_invoice(
            amount_sats=int(total_amount),
            memo=memo,
            expires_in_minutes=30  # 결제용 인보이스는 30분 유효
        )
        
        if settings.DEBUG:
            logger.debug(f"[INVOICE] 인보이스 생성 결과: {result}")
        
        if not result['success']:
            return JsonResponse({
                'success': False,
                'error': result['error']
            }, status=500)
        
        # 인보이스 데이터베이스에 저장
        try:
            # 사용자가 로그인되어 있지 않으면 None으로 저장
            user_for_invoice = request.user if request.user.is_authenticated else None
            
            # 🛡️ 기존 pending 상태의 인보이스 초기화 (재생성 대비)
            try:
                if user_for_invoice:
                    # 로그인된 사용자의 기존 pending 인보이스 취소
                    existing_invoices = Invoice.objects.filter(
                        user=user_for_invoice,
                        store=first_store,
                        status='pending'
                    )
                else:
                    # 비로그인 사용자의 경우 현재 세션 기반으로 처리하기 어려우므로 
                    # 같은 스토어의 최근 pending 인보이스들 중 만료된 것들을 정리
                    from datetime import timedelta
                    cutoff_time = timezone.now() - timedelta(hours=1)  # 1시간 이전 것들 정리
                    existing_invoices = Invoice.objects.filter(
                        user=None,
                        store=first_store,
                        status='pending',
                        created_at__lt=cutoff_time
                    )
                
                if existing_invoices.exists():
                    existing_invoices.update(status='cancelled')
                    if settings.DEBUG:
                        logger.debug(f"[INVOICE] 기존 pending 인보이스 {existing_invoices.count()}개 취소됨")
                        
            except Exception as e:
                if settings.DEBUG:
                    logger.warning(f"[INVOICE] 기존 인보이스 정리 실패: {str(e)}")
                # 실패해도 계속 진행
            
            invoice = Invoice.objects.create(
                payment_hash=result['payment_hash'],
                invoice_string=result['invoice'],
                amount_sats=int(total_amount),
                memo=memo,
                user=user_for_invoice,
                store=first_store,
                expires_at=result['expires_at'],
                status='pending'
            )
            
            if settings.DEBUG:
                logger.debug(f"[INVOICE] 인보이스 DB 저장 완료: ID={invoice.id}")
                
        except Exception as e:
            if settings.DEBUG:
                logger.error(f"[INVOICE] 인보이스 DB 저장 실패: {str(e)}")
            # DB 저장 실패해도 인보이스는 생성되었으므로 계속 진행
        
        # 응답 데이터 준비
        cart_summary = cart_service.get_cart_summary()
        response_data = {
            'success': True,
            'payment_hash': result['payment_hash'],
            'invoice': result['invoice'],
            'amount': int(total_amount),
            'memo': memo,
            'expires_at': result['expires_at'].isoformat() if result.get('expires_at') else None,
            'stores_count': len(stores_with_items),
            'items_count': cart_summary['total_items']
        }
        
        if settings.DEBUG:
            logger.debug(f"응답 데이터: {response_data}")
        return JsonResponse(response_data)
        
    except json.JSONDecodeError as e:
        if settings.DEBUG:
            logger.error(f"JSON 디코딩 오류: {str(e)}")
        return JsonResponse({'success': False, 'error': f'JSON 파싱 오류: {str(e)}'}, status=400)
    except Exception as e:
        if settings.DEBUG:
            logger.error(f"create_checkout_invoice 예외 발생: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'인보이스 생성 중 오류가 발생했습니다: {str(e)}'}, status=500)


@require_POST
def check_checkout_payment(request):
    """결제 상태 확인 API"""
    try:
        if settings.DEBUG:
            logger.debug("check_checkout_payment 호출")
        
        # 요청 본문 파싱
        try:
            data = json.loads(request.body)
            if settings.DEBUG:
                logger.debug(f"결제 상태 확인 요청 데이터: {data}")
        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'error': f'JSON 파싱 오류: {str(e)}'}, status=400)
        
        payment_hash = data.get('payment_hash')
        if not payment_hash:
            return JsonResponse({'success': False, 'error': 'payment_hash가 필요합니다.'}, status=400)
        
        if settings.DEBUG:
            logger.debug(f"결제 상태 확인: payment_hash={payment_hash}")
        
        # 🛡️ 이미 결제 완료된 주문이 있는지 먼저 확인 (중복 이메일 발송 방지)
        existing_orders = Order.objects.filter(payment_id=payment_hash)
        if existing_orders.exists():
            if settings.DEBUG:
                logger.debug(f"[PAYMENT] 이미 결제 완료된 주문 발견: {payment_hash}")
            
            all_orders = list(existing_orders)
            return JsonResponse({
                'success': True,
                'status': 'paid',
                'paid': True,
                'order_number': all_orders[0].order_number if all_orders else None,
                'redirect_url': f'/orders/checkout/complete/{all_orders[0].order_number}/' if all_orders else None
            })
        
        # 장바구니 확인
        cart_service = CartService(request)
        cart_items = cart_service.get_cart_items()
        
        if not cart_items:
            return JsonResponse({'success': False, 'error': '장바구니가 비어있습니다.'}, status=400)
        
        # 첫 번째 스토어 가져오기
        first_item = cart_items[0]
        try:
            first_store = Store.objects.filter(store_id=first_item['store_id'], deleted_at__isnull=True).first()
            if not first_store:
                return JsonResponse({'success': False, 'error': '스토어를 찾을 수 없습니다.'}, status=400)
        except Exception:
            return JsonResponse({'success': False, 'error': '스토어를 찾을 수 없습니다.'}, status=400)
        
        # BlinkAPIService 초기화
        try:
            blink_service = get_blink_service_for_store(first_store)
            if settings.DEBUG:
                logger.debug("스토어 정보로 BlinkAPIService 초기화")
        except Exception as e:
            if settings.DEBUG:
                logger.error(f"BlinkAPIService 초기화 실패: {str(e)}")
            return JsonResponse({'success': False, 'error': f'Blink API 서비스 초기화 실패: {str(e)}'}, status=500)
        
        # 결제 상태 확인
        result = blink_service.check_invoice_status(payment_hash)
        
        if settings.DEBUG:
            logger.debug(f"결제 상태 확인 결과: {result}")
        
        if not result['success']:
            return JsonResponse({
                'success': False,
                'error': result['error']
            })
        
        # 결제 완료 시 주문 생성
        if result['status'] == 'paid':
            try:
                # 🛡️ 트랜잭션 및 select_for_update로 동시성 문제 방지
                with transaction.atomic():
                    # 다시 한 번 중복 결제 확인 (트랜잭션 내에서)
                    existing_orders = Order.objects.filter(payment_id=payment_hash)
                    if existing_orders.exists():
                        if settings.DEBUG:
                            logger.debug(f"[PAYMENT] 트랜잭션 내 중복 결제 확인: {payment_hash}")
                        
                        all_orders = list(existing_orders)
                        return JsonResponse({
                            'success': True,
                            'status': result['status'],
                            'paid': True,
                            'order_number': all_orders[0].order_number if all_orders else None,
                            'redirect_url': f'/orders/checkout/complete/{all_orders[0].order_number}/' if all_orders else None
                        })
                    
                    # 인보이스 상태 업데이트
                    try:
                        invoice = Invoice.objects.select_for_update().get(payment_hash=payment_hash)
                        if invoice.status == 'paid':
                            # 이미 처리된 인보이스인 경우
                            if settings.DEBUG:
                                logger.debug(f"[PAYMENT] 이미 처리된 인보이스: {payment_hash}")
                            return JsonResponse({
                                'success': True,
                                'status': result['status'],
                                'paid': True,
                                'order_number': invoice.order.order_number if invoice.order else None,
                                'redirect_url': f'/orders/checkout/complete/{invoice.order.order_number}/' if invoice.order else None
                            })
                        
                        invoice.status = 'paid'
                        invoice.paid_at = timezone.now()
                        invoice.save()
                        
                        if settings.DEBUG:
                            logger.debug(f"[PAYMENT] 인보이스 상태 업데이트 완료: {payment_hash}")
                    except Invoice.DoesNotExist:
                        if settings.DEBUG:
                            logger.warning(f"[PAYMENT] 인보이스를 찾을 수 없음: {payment_hash}")
                    
                    # 주문 생성 로직 (새로운 CartService 기반 함수 사용)
                    shipping_data = request.session.get('shipping_data', {})
                    order_result = create_order_from_cart_service(request, payment_hash, shipping_data)
                    
                    # 인보이스와 주문 연결 (첫 번째 주문과 연결)
                    try:
                        invoice = Invoice.objects.get(payment_hash=payment_hash)
                        if order_result['orders']:
                            primary_order = order_result['orders'][0]
                            invoice.order = primary_order
                            invoice.save()
                            
                            if settings.DEBUG:
                                logger.debug(f"[PAYMENT] 인보이스-주문 연결 완료: {payment_hash} -> {primary_order.order_number}")
                    except (Invoice.DoesNotExist, Order.DoesNotExist) as e:
                        if settings.DEBUG:
                            logger.warning(f"[PAYMENT] 인보이스-주문 연결 실패: {str(e)}")
                    
                    # 주문 완료 후 세션에서 배송 정보 삭제
                    if 'shipping_data' in request.session:
                        del request.session['shipping_data']
                    
                    return JsonResponse({
                        'success': True,
                        'status': result['status'],
                        'paid': True,
                        'order_number': order_result['primary_order_number'],
                        'redirect_url': f'/orders/checkout/complete/{order_result["primary_order_number"]}/'
                    })
                    
            except Exception as e:
                if settings.DEBUG:
                    logger.error(f"주문 생성 실패: {str(e)}", exc_info=True)
                
                # 주문 생성 실패 시 인보이스 상태를 다시 pending으로 변경
                try:
                    invoice = Invoice.objects.get(payment_hash=payment_hash)
                    invoice.status = 'pending'
                    invoice.paid_at = None
                    invoice.save()
                except Invoice.DoesNotExist:
                    pass
                
                return JsonResponse({
                    'success': False,
                    'error': f'주문 생성 중 오류가 발생했습니다: {str(e)}'
                }, status=500)
        
        return JsonResponse({
            'success': True,
            'status': result['status'],
            'paid': False
        })
        
    except Exception as e:
        if settings.DEBUG:
            logger.error(f"check_checkout_payment 예외 발생: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'결제 상태 확인 중 오류가 발생했습니다: {str(e)}'}, status=500)


@require_POST
def cancel_invoice(request):
    """인보이스 취소"""
    try:
        if settings.DEBUG:
            logger.debug("cancel_invoice 호출")
        
        # 요청 본문 파싱
        try:
            data = json.loads(request.body)
            if settings.DEBUG:
                logger.debug(f"인보이스 취소 요청 데이터: {data}")
        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'error': f'JSON 파싱 오류: {str(e)}'}, status=400)
        
        payment_hash = data.get('payment_hash')
        if not payment_hash:
            return JsonResponse({'success': False, 'error': 'payment_hash가 필요합니다.'}, status=400)
        
        # 🛡️ 이미 결제 완료된 주문이 있는지 먼저 확인 (취소 불가)
        existing_orders = Order.objects.filter(payment_id=payment_hash)
        if existing_orders.exists():
            if settings.DEBUG:
                logger.debug(f"[CANCEL] 이미 결제 완료된 주문 발견: {payment_hash}")
            
            all_orders = list(existing_orders)
            return JsonResponse({
                'success': False,
                'error': '이미 결제가 완료되었습니다. 취소할 수 없습니다.',
                'redirect_url': f'/orders/checkout/complete/{all_orders[0].order_number}/' if all_orders else None
            })
        
        # 인보이스 찾기 (로그인/비로그인 사용자 모두 지원)
        try:
            if request.user.is_authenticated:
                # 로그인된 사용자의 경우 사용자별 인보이스 찾기
                invoice = Invoice.objects.get(payment_hash=payment_hash, user=request.user)
            else:
                # 비로그인 사용자의 경우 user가 None인 인보이스 찾기
                invoice = Invoice.objects.get(payment_hash=payment_hash, user=None)
        except Invoice.DoesNotExist:
            return JsonResponse({'success': False, 'error': '인보이스를 찾을 수 없습니다.'}, status=404)
        
        # 취소 가능한 상태인지 확인
        if invoice.status != 'pending':
            return JsonResponse({
                'success': False, 
                'error': f'취소할 수 없는 상태입니다. 현재 상태: {invoice.get_status_display()}'
            }, status=400)
        
        # 🛡️ 실제 결제 상태를 Blink API로 재확인
        try:
            # 스토어 가져오기
            store = invoice.store
            
            # BlinkAPIService 초기화
            blink_service = get_blink_service_for_store(store)
            
            # 결제 상태 재확인
            result = blink_service.check_invoice_status(payment_hash)
            
            if result['success'] and result['status'] == 'paid':
                # 실제로는 결제가 완료되었음!
                if settings.DEBUG:
                    logger.debug(f"[CANCEL] 실제 결제 완료 감지: {payment_hash}")
                
                # 인보이스 상태 업데이트
                invoice.status = 'paid'
                invoice.paid_at = timezone.now()
                invoice.save()
                
                # 주문 생성 (결제 완료 처리)
                try:
                    with transaction.atomic():
                        # 주문 생성 로직
                        shipping_data = request.session.get('shipping_data', {})
                        order_result = create_order_from_cart_service(request, payment_hash, shipping_data)
                        
                        # 인보이스와 주문 연결
                        if order_result['orders']:
                            primary_order = order_result['orders'][0]
                            invoice.order = primary_order
                            invoice.save()
                        
                        # 배송 정보 삭제
                        if 'shipping_data' in request.session:
                            del request.session['shipping_data']
                        
                        return JsonResponse({
                            'success': False,
                            'error': '결제가 완료되었습니다. 주문 완료 페이지로 이동합니다.',
                            'redirect_url': f'/orders/checkout/complete/{order_result["primary_order_number"]}/'
                        })
                        
                except Exception as e:
                    if settings.DEBUG:
                        logger.error(f"[CANCEL] 주문 생성 실패: {str(e)}")
                    
                    return JsonResponse({
                        'success': False,
                        'error': '결제는 완료되었지만 주문 생성에 실패했습니다. 고객센터에 문의해주세요.'
                    })
            
        except Exception as e:
            if settings.DEBUG:
                logger.warning(f"[CANCEL] 결제 상태 재확인 실패: {str(e)}")
            # 재확인 실패 시에는 그대로 진행
        
        # 인보이스 상태를 취소로 변경
        invoice.status = 'cancelled'
        invoice.save()
        
        if settings.DEBUG:
            logger.debug(f"인보이스 취소 완료: {payment_hash}")
        
        return JsonResponse({
            'success': True,
            'message': '인보이스가 취소되었습니다.'
        })
        
    except Exception as e:
        if settings.DEBUG:
            logger.error(f"cancel_invoice 예외 발생: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'인보이스 취소 중 오류가 발생했습니다: {str(e)}'}, status=500)


def create_order_from_cart_service(request, payment_hash, shipping_data=None):
    """CartService를 사용하여 장바구니에서 주문 생성 (로그인/비로그인 모두 지원)"""
    import uuid
    
    # 🛡️ 중복 주문 생성 방지: 이미 해당 payment_hash로 주문이 존재하는지 확인
    existing_orders = Order.objects.filter(payment_id=payment_hash)
    if existing_orders.exists():
        if settings.DEBUG:
            logger.debug(f"[ORDER_CREATE] 중복 주문 생성 방지: {payment_hash} - 기존 주문 {existing_orders.count()}개 발견")
        
        # 기존 주문 정보 반환
        all_orders = list(existing_orders)
        return {
            'orders': all_orders,
            'primary_order_number': all_orders[0].order_number if all_orders else None,
            'total_orders': len(all_orders),
            'total_amount': sum(order.total_amount for order in all_orders),
            'total_subtotal': sum(order.subtotal for order in all_orders),
            'total_shipping_fee': sum(order.shipping_fee for order in all_orders)
        }
    
    cart_service = CartService(request)
    cart_items = cart_service.get_cart_items()
    
    if not cart_items:
        raise Exception('장바구니가 비어있습니다.')
    
    user = request.user if request.user.is_authenticated else None
    
    # 비회원 주문인 경우 기본 사용자 정보 설정
    if not user:
        if not shipping_data:
            raise Exception('비회원 주문 시 배송 정보가 필요합니다.')
        
        # 비회원 주문의 경우 익명 사용자 생성 또는 특별 처리
        # 여기서는 None으로 두고 Order 모델에서 nullable로 처리하거나
        # 시스템 사용자를 사용합니다
        from django.contrib.auth.models import User
        try:
            # 시스템 사용자 가져오기 또는 생성
            user, created = User.objects.get_or_create(
                username='anonymous_guest',
                defaults={
                    'email': 'anonymous@satoshop.com',
                    'first_name': 'Anonymous',
                    'last_name': 'Guest'
                }
            )
        except Exception:
            raise Exception('비회원 주문 처리 중 오류가 발생했습니다.')
    
    # 스토어별로 그룹화
    stores_data = {}
    for item in cart_items:
        store_id = item['store_id']
        if store_id not in stores_data:
            stores_data[store_id] = {
                'store_name': item['store_name'],
                'store_id': store_id,
                'items': []
            }
        stores_data[store_id]['items'].append(item)
    
    stores_orders = {}
    all_orders = []
    
    # 배송 정보 기본값 설정
    if not shipping_data:
        shipping_data = {
            'buyer_name': user.get_full_name() or user.username,
            'buyer_phone': '',
            'buyer_email': user.email,
            'shipping_postal_code': '',
            'shipping_address': '디지털 상품',
            'shipping_detail_address': '',
            'order_memo': '라이트닝 결제로 주문',
        }
    
    with transaction.atomic():
        for store_id, store_data in stores_data.items():
            store_items = store_data['items']
            store = Store.objects.filter(store_id=store_id, deleted_at__isnull=True).first()
            if not store:
                continue
            
            if settings.DEBUG:
                logger.debug(f"[ORDER_CREATE] 스토어: {store.store_name}, 상품 수: {len(store_items)}")
            
            # 주문 번호는 Order 모델의 save() 메소드에서 자동 생성됨
            
            store_subtotal, store_shipping_fee, override_free = calculate_store_totals(store, store_items)

            # 주문 생성
            order = Order.objects.create(
                user=user,
                store=store,
                buyer_name=shipping_data.get('buyer_name', ''),
                buyer_phone=shipping_data.get('buyer_phone', ''),
                buyer_email=shipping_data.get('buyer_email', ''),
                shipping_postal_code=shipping_data.get('shipping_postal_code', ''),
                shipping_address=shipping_data.get('shipping_address', ''),
                shipping_detail_address=shipping_data.get('shipping_detail_address', ''),
                order_memo=shipping_data.get('order_memo', ''),
                subtotal=0,  # 나중에 계산
                shipping_fee=store_shipping_fee,
                total_amount=0,  # 나중에 계산
                status='paid',
                payment_id=payment_hash,
                paid_at=timezone.now()
            )
            
            order.force_free_override = override_free
            stores_orders[store.id] = order
            all_orders.append(order)
            
            if settings.DEBUG:
                logger.debug(f"[ORDER_CREATE] 주문 생성: {order.order_number}, 배송비: {store_shipping_fee} sats")
            
            # 주문 아이템 생성
            for item in store_items:
                try:
                    product = Product.objects.get(id=item['product_id'])
                except Product.DoesNotExist:
                    continue
                
                # 재고 확인 및 감소
                if product.is_temporarily_out_of_stock:
                    raise Exception(f'"{product.title}" 상품이 일시품절 상태입니다.')
                elif not product.can_purchase(item['quantity']):
                    raise Exception(f'"{product.title}" 상품의 재고가 부족합니다. (요청: {item["quantity"]}개, 재고: {product.stock_quantity}개)')
                
                # 재고 감소
                if not product.decrease_stock(item['quantity']):
                    raise Exception(f'"{product.title}" 상품의 재고 감소에 실패했습니다.')
                
                # 옵션 정보를 문자열 형태로 변환
                options_display = {}
                if item.get('selected_options'):
                    for option_id, choice_id in item['selected_options'].items():
                        try:
                            from products.models import ProductOption, ProductOptionChoice
                            option = ProductOption.objects.get(id=option_id)
                            choice = ProductOptionChoice.objects.get(id=choice_id)
                            options_display[option.name] = choice.name
                        except (ProductOption.DoesNotExist, ProductOptionChoice.DoesNotExist):
                            continue
                
                # 장바구니의 고정된 가격 사용 (환율 고정 반영)
                unit_price = item['unit_price']  # 이미 고정된 가격 포함
                
                # 기존 로직 호환을 위해 상품 가격과 옵션 가격 분리
                # 하지만 총합은 장바구니의 고정된 가격을 사용
                # 할인 상품인 경우 할인가를 사용
                if hasattr(item, 'frozen_product_price_sats') and item.get('frozen_product_price_sats'):
                    base_product_price = item.get('frozen_product_price_sats')
                elif product.is_discounted and product.public_discounted_price:
                    base_product_price = product.public_discounted_price
                else:
                    base_product_price = product.public_price
                
                options_price = unit_price - base_product_price if unit_price > base_product_price else 0
                
                order_item = OrderItem.objects.create(
                    order=order,
                    product=product,
                    product_title=product.title,
                    product_price=base_product_price,
                    quantity=item['quantity'],
                    selected_options=options_display,
                    options_price=options_price
                )
                
                if settings.DEBUG:
                    logger.debug(f"[ORDER_CREATE]   - 상품: {product.title}, 수량: {item['quantity']}, 총가격: {order_item.total_price} sats, 남은재고: {product.stock_quantity}개")
        
        # 각 주문의 총액 계산
        for order in stores_orders.values():
            order.subtotal = sum(item.total_price for item in order.items.all())
            order.total_amount = order.subtotal + order.shipping_fee
            order.save()
            
            if settings.DEBUG:
                logger.debug(f"[ORDER_CREATE] 주문 총액 계산: {order.order_number}, 상품: {order.subtotal} + 배송: {order.shipping_fee} = {order.total_amount} sats")
            
            # 구매 내역 생성
            if user:
                PurchaseHistory.objects.create(
                    user=user,
                    order=order,
                    store_name=order.store.store_name,
                    total_amount=order.total_amount,
                    purchase_date=order.paid_at
                )
            
        # 🎉 주문 완료 이메일 발송 (스토어별로 중복 방지)
        email_sent_stores = set()  # 이메일 발송한 스토어 추적
        
        for order in stores_orders.values():
            # 이미 이메일을 발송한 스토어는 건너뛰기
            if order.store.id in email_sent_stores:
                if settings.DEBUG:
                    logger.debug(f"[ORDER_EMAIL] 스토어 {order.store.store_name}에 이미 이메일 발송됨, 건너뛰기: {order.order_number}")
                continue
            
            try:
                from .services import send_order_notification_email
                email_sent = send_order_notification_email(order)
                if email_sent:
                    email_sent_stores.add(order.store.id)  # 발송 완료 스토어 기록
                    if settings.DEBUG:
                        logger.debug(f"[ORDER_EMAIL] 주문 알림 이메일 발송 성공: {order.order_number}")
                else:
                    if settings.DEBUG:
                        logger.debug(f"[ORDER_EMAIL] 주문 알림 이메일 발송 조건 미충족: {order.order_number}")
            except Exception as e:
                # 이메일 발송 실패해도 주문 처리는 계속 진행
                logger.error(f"[ORDER_EMAIL] 주문 알림 이메일 발송 오류: {order.order_number}, {str(e)}")
                pass
        
        # 장바구니 비우기
        cart_service.clear_cart()
        
        # 모든 주문 정보 반환
        total_amount = sum(order.total_amount for order in all_orders)
        total_subtotal = sum(order.subtotal for order in all_orders)
        total_shipping_fee = sum(order.shipping_fee for order in all_orders)
        
        if settings.DEBUG:
            logger.debug(f"[ORDER_CREATE] 전체 주문 완료: {len(all_orders)}개, 총액: {total_amount} sats")
        
        return {
            'orders': all_orders,
            'primary_order_number': all_orders[0].order_number if all_orders else None,
            'total_orders': len(all_orders),
            'total_amount': total_amount,
            'total_subtotal': total_subtotal,
            'total_shipping_fee': total_shipping_fee
        }


def create_order_from_cart(user, cart, payment_hash, shipping_data=None):
    """장바구니에서 주문 생성 (기존 함수 - 호환성 유지)"""
    import uuid
    from itertools import groupby
    
    # 스토어별로 정렬한 후 그룹화 (중요: groupby는 정렬된 데이터에서만 올바르게 작동)
    cart_items = cart.items.all().select_related('product', 'product__store').order_by('product__store__id')
    
    # 스토어별로 그룹화
    stores_orders = {}
    all_orders = []
    
    # 배송 정보 기본값 설정
    if not shipping_data:
        shipping_data = {
            'buyer_name': user.get_full_name() or user.username,
            'buyer_phone': '',
            'buyer_email': user.email,
            'shipping_postal_code': '',
            'shipping_address': '디지털 상품',
            'shipping_detail_address': '',
            'order_memo': '라이트닝 결제로 주문',
        }
    
    with transaction.atomic():
        for store, items in groupby(cart_items, key=lambda x: x.product.store):
            store_items = list(items)

            if settings.DEBUG:
                logger.debug(f"[ORDER_CREATE] 스토어: {store.store_name}, 상품 수: {len(store_items)}")

            # 주문 번호는 Order 모델의 save() 메소드에서 자동 생성됨

            # 스토어별 배송비 계산 (스토어 정책 기준)
            store_subtotal, store_shipping_fee, override_free = calculate_store_totals(store, store_items)

            # 주문 생성
            order = Order.objects.create(
                user=user,
                store=store,
                buyer_name=shipping_data.get('buyer_name', ''),
                buyer_phone=shipping_data.get('buyer_phone', ''),
                buyer_email=shipping_data.get('buyer_email', ''),
                shipping_postal_code=shipping_data.get('shipping_postal_code', ''),
                shipping_address=shipping_data.get('shipping_address', ''),
                shipping_detail_address=shipping_data.get('shipping_detail_address', ''),
                order_memo=shipping_data.get('order_memo', ''),
                subtotal=0,  # 나중에 계산
                shipping_fee=store_shipping_fee,
                total_amount=0,  # 나중에 계산
                status='paid',
                payment_id=payment_hash,
                paid_at=timezone.now()
            )
            
            order.force_free_override = override_free
            stores_orders[store.id] = order
            all_orders.append(order)
            
            if settings.DEBUG:
                logger.debug(f"[ORDER_CREATE] 주문 생성: {order.order_number}, 배송비: {store_shipping_fee} sats")
            
            # 주문 아이템 생성
            for item in store_items:
                # 재고 확인 및 감소
                if item.product.is_temporarily_out_of_stock:
                    raise Exception(f'"{item.product.title}" 상품이 일시품절 상태입니다.')
                elif not item.product.can_purchase(item.quantity):
                    raise Exception(f'"{item.product.title}" 상품의 재고가 부족합니다. (요청: {item.quantity}개, 재고: {item.product.stock_quantity}개)')
                
                # 재고 감소
                if not item.product.decrease_stock(item.quantity):
                    raise Exception(f'"{item.product.title}" 상품의 재고 감소에 실패했습니다.')
                
                # 옵션 정보를 문자열 형태로 변환
                options_display = {}
                for option_id, choice_id in item.selected_options.items():
                    try:
                        from products.models import ProductOption, ProductOptionChoice
                        option = ProductOption.objects.get(id=option_id)
                        choice = ProductOptionChoice.objects.get(id=choice_id)
                        options_display[option.name] = choice.name
                    except (ProductOption.DoesNotExist, ProductOptionChoice.DoesNotExist):
                        continue
                
                # 할인 상품인 경우 할인가를 사용
                if item.product.is_discounted and item.product.public_discounted_price:
                    product_price = item.product.public_discounted_price
                else:
                    product_price = item.product.public_price
                
                order_item = OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    product_title=item.product.title,
                    product_price=product_price,
                    quantity=item.quantity,
                    selected_options=options_display,
                    options_price=item.options_price
                )
                
                if settings.DEBUG:
                    logger.debug(f"[ORDER_CREATE]   - 상품: {item.product.title}, 수량: {item.quantity}, 총가격: {order_item.total_price} sats, 남은재고: {item.product.stock_quantity}개")
        
        # 각 주문의 총액 계산
        for order in stores_orders.values():
            order.subtotal = sum(item.total_price for item in order.items.all())
            order.total_amount = order.subtotal + order.shipping_fee
            order.save()
            
            if settings.DEBUG:
                logger.debug(f"[ORDER_CREATE] 주문 총액 계산: {order.order_number}, 상품: {order.subtotal} + 배송: {order.shipping_fee} = {order.total_amount} sats")
            
            # 구매 내역 생성
            PurchaseHistory.objects.create(
                user=user,
                order=order,
                store_name=order.store.store_name,
                total_amount=order.total_amount,
                purchase_date=order.paid_at
            )
        
        # 장바구니 비우기
        cart.items.all().delete()
        
        # 모든 주문 정보 반환
        total_amount = sum(order.total_amount for order in all_orders)
        total_subtotal = sum(order.subtotal for order in all_orders)
        total_shipping_fee = sum(order.shipping_fee for order in all_orders)
        
        if settings.DEBUG:
            logger.debug(f"[ORDER_CREATE] 전체 주문 완료: {len(all_orders)}개, 총액: {total_amount} sats")
        
        return {
            'orders': all_orders,
            'primary_order_number': all_orders[0].order_number if all_orders else None,
            'total_orders': len(all_orders),
            'total_amount': total_amount,
            'total_subtotal': total_subtotal,
            'total_shipping_fee': total_shipping_fee
        }


def download_order_txt_public(request, order_number):
    """주문서 TXT 파일 다운로드 (공개 - 주문번호만으로 접근 가능)"""
    try:
        order = get_object_or_404(Order, order_number=order_number)
    except Order.DoesNotExist:
        return HttpResponse("주문을 찾을 수 없습니다.", status=404)
    
    # 추가 보안 체크: 최근 24시간 내 주문만 허용 (선택적)
    # from datetime import timedelta
    # if timezone.now() - order.created_at > timedelta(hours=24):
    #     return HttpResponse("다운로드 기간이 만료되었습니다.", status=403)
    
    # TXT 내용 생성 (새로운 포맷터 사용)
    from .formatters import generate_txt_order
    content = generate_txt_order(order)

    # HTTP 응답 생성 (BOM 추가로 인코딩 문제 해결)
    content_with_bom = '\ufeff' + content  # UTF-8 BOM 추가
    response = HttpResponse(content_with_bom, content_type='text/plain; charset=utf-8')
    
    # 파일명 인코딩 처리 (모바일 브라우저 호환성 개선)
    filename = f"주문서_{order.order_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.txt"
    try:
        # RFC 5987 방식으로 UTF-8 파일명 인코딩
        from urllib.parse import quote
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
    except:
        # 백업 방식: 영문 파일명 사용
        fallback_filename = f"order_{order.order_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.txt"
        response['Content-Disposition'] = f'attachment; filename="{fallback_filename}"'
    
    return response


@login_required
def toggle_delivery_status(request, order_id):
    """배송상태 토글 API"""
    if request.method != 'POST':
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=405)
    
    try:
        # 주문 조회
        order = Order.objects.get(id=order_id)
        
        # 권한 확인 (스토어 주인만 변경 가능)
        if order.store.owner != request.user:
            return JsonResponse({'error': '권한이 없습니다.'}, status=403)
        
        # 배송상태 토글
        if order.delivery_status == 'preparing':
            order.delivery_status = 'completed'
        else:
            order.delivery_status = 'preparing'
        
        order.save()
        
        return JsonResponse({
            'success': True,
            'delivery_status': order.delivery_status,
            'delivery_status_display': order.get_delivery_status_display()
        })
        
    except Order.DoesNotExist:
        return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def update_tracking_info(request, order_id):
    """택배 정보 업데이트 API"""
    if request.method != 'POST':
        return JsonResponse({'error': '잘못된 요청입니다.'}, status=405)
    
    try:
        # JSON 데이터 파싱
        data = json.loads(request.body)
        courier_company = data.get('courier_company', '').strip()
        tracking_number = data.get('tracking_number', '').strip()
        
        # 주문 조회
        order = Order.objects.get(id=order_id)
        
        # 권한 확인 (스토어 주인만 변경 가능)
        if order.store.owner != request.user:
            return JsonResponse({'error': '권한이 없습니다.'}, status=403)
        
        # 택배 정보 업데이트
        order.courier_company = courier_company
        order.tracking_number = tracking_number
        order.save()
        
        return JsonResponse({
            'success': True,
            'courier_company': order.courier_company,
            'tracking_number': order.tracking_number
        })
        
    except Order.DoesNotExist:
        return JsonResponse({'error': '주문을 찾을 수 없습니다.'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 JSON 형식입니다.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
