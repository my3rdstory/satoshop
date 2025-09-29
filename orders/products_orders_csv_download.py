import csv
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from .models import OrderItem


def _format_text_for_csv(value):
    """CSV 다운로드 시 Excel에서 일반 텍스트로 유지되도록 포맷팅"""
    if value in (None, ''):
        return ''

    text_value = str(value)
    escaped = text_value.replace('"', '""')
    return f'="{escaped}"'

def export_product_orders_csv(request, store, product):
    """
    현재 화면에 표시된 상품 주문 데이터를 CSV로 다운로드
    화면의 필터 상태(전체/이번달/지난달/기간선택)를 그대로 반영
    """
    
    # 현재 날짜 정보
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    
    # 이번달 범위
    current_month_start = timezone.datetime(current_year, current_month, 1, tzinfo=timezone.get_current_timezone())
    if current_month == 12:
        current_month_end = timezone.datetime(current_year + 1, 1, 1, tzinfo=timezone.get_current_timezone())
    else:
        current_month_end = timezone.datetime(current_year, current_month + 1, 1, tzinfo=timezone.get_current_timezone())
    
    # 지난달 범위
    last_month_date = now - relativedelta(months=1)
    last_month_year = last_month_date.year
    last_month_month = last_month_date.month
    last_month_start = timezone.datetime(last_month_year, last_month_month, 1, tzinfo=timezone.get_current_timezone())
    if last_month_month == 12:
        last_month_end = timezone.datetime(last_month_year + 1, 1, 1, tzinfo=timezone.get_current_timezone())
    else:
        last_month_end = timezone.datetime(last_month_year, last_month_month + 1, 1, tzinfo=timezone.get_current_timezone())
    
    # URL 파라미터에서 필터 정보 가져오기
    filter_type = request.GET.get('filter', 'this_month')  # all, this_month, last_month, custom
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    page_number = request.GET.get('page', 1)
    
    # 디버깅: URL 파라미터 로그
    print(f"CSV 다운로드 파라미터: filter={filter_type}, start_date={start_date}, end_date={end_date}, page={page_number}")
    print(f"전체 GET 파라미터: {dict(request.GET)}")
    
    # 기본 쿼리셋
    order_items = OrderItem.objects.filter(
        product=product
    ).select_related('order')
    
    # 날짜 필터링 적용 (화면과 동일한 로직)
    if filter_type == 'this_month':
        order_items = order_items.filter(
            order__created_at__gte=current_month_start,
            order__created_at__lt=current_month_end
        )
        filter_name = f"{current_month}월"
        print(f"이번달 필터 적용: {current_month_start} ~ {current_month_end}")
    elif filter_type == 'last_month':
        order_items = order_items.filter(
            order__created_at__gte=last_month_start,
            order__created_at__lt=last_month_end
        )
        filter_name = f"{last_month_month}월"
        print(f"지난달 필터 적용: {last_month_start} ~ {last_month_end}")
    elif filter_type == 'custom' and start_date and end_date:
        try:
            start_datetime = timezone.datetime.strptime(start_date, '%Y-%m-%d')
            start_datetime = start_datetime.replace(tzinfo=timezone.get_current_timezone())
            end_datetime = timezone.datetime.strptime(end_date, '%Y-%m-%d')
            end_datetime = end_datetime.replace(tzinfo=timezone.get_current_timezone()) + timedelta(days=1)  # 종료일 포함
            order_items = order_items.filter(
                order__created_at__gte=start_datetime,
                order__created_at__lt=end_datetime
            )
            filter_name = f"{start_date}~{end_date}"
            print(f"커스텀 필터 적용: {start_datetime} ~ {end_datetime}")
        except ValueError:
            # 날짜 형식이 잘못된 경우 전체 조회
            filter_name = "전체"
            print("커스텀 필터 날짜 파싱 실패, 전체 조회")
    elif filter_type == 'all':
        # 전체 필터 - 날짜 제한 없음
        filter_name = "전체"
        print("전체 필터 적용")
    else:
        # 필터 타입이 없거나 알 수 없는 경우 - 전체 조회
        filter_name = "전체"
        print(f"알 수 없는 필터 타입 또는 파라미터 없음: {filter_type}, 전체 조회로 처리")
    
    # 정렬 적용 (최신순)
    order_items = order_items.order_by('-order__created_at')
    
    print(f"필터링 후 총 주문 수: {order_items.count()}")
    
    # 페이지네이션이 적용된 경우 해당 페이지의 데이터만 가져오기 (전체 필터일 때만)
    if filter_type == 'all' and page_number:
        print(f"전체 필터 페이지네이션 적용: 페이지 {page_number}")
        paginator = Paginator(order_items, 10)
        try:
            page_obj = paginator.get_page(page_number)
            order_items = page_obj.object_list
            filter_name += f"_페이지{page_number}"
            print(f"페이지네이션 후 주문 수: {len(order_items)}")
        except:
            # 페이지 번호가 잘못된 경우 첫 페이지
            page_obj = paginator.get_page(1)
            order_items = page_obj.object_list
            filter_name += "_페이지1"
            print(f"페이지네이션 실패, 첫 페이지로 처리. 주문 수: {len(order_items)}")
    else:
        print(f"페이지네이션 없음. 전체 데이터 다운로드. 주문 수: {order_items.count()}")
    
    # CSV 응답 생성
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    
    # 파일명에 필터 정보 포함
    filename = f"{store.store_id}_{product.title}_{filter_name}_주문목록.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # UTF-8 BOM 추가 (엑셀에서 한글이 깨지지 않도록)
    response.write('\ufeff')
    
    # CSV writer 생성
    writer = csv.writer(response)
    
    # 헤더 작성
    writer.writerow([
        '주문번호', '주문자', '수량', '단가(sats)', '옵션가격(sats)', '총액(sats)', 
        '주문일시', '상태', '배송상태', '선택옵션', '연락처', '이메일', 
        '우편번호', '주소', '상세주소', '요청사항'
    ])
    
    # 데이터 작성
    for item in order_items:
        order = item.order
        
        # 선택된 옵션을 문자열로 변환
        options_str = ''
        if item.selected_options:
            options_list = []
            for key, value in item.selected_options.items():
                options_list.append(f"{key}: {value}")
            options_str = ', '.join(options_list)
        
        # 우편번호 처리 - 0으로 시작하는 경우 문자열로 보존
        postal_code = _format_text_for_csv(order.shipping_postal_code or '')
        phone_value = _format_text_for_csv(order.buyer_phone or '')
        
        writer.writerow([
            order.order_number,
            order.buyer_name,
            item.quantity,
            item.product_price,
            item.options_price,
            item.total_price,
            timezone.localtime(order.created_at).strftime('%Y-%m-%d %H:%M:%S'),
            order.get_status_display(),
            order.get_delivery_status_display(),
            options_str,
            phone_value,
            order.buyer_email or '',
            postal_code,
            order.shipping_address or '',
            order.shipping_detail_address or '',
            order.order_memo or '',
        ])
    
    return response


def export_product_orders_excel(request, store, product):
    """
    현재 화면에 표시된 상품 주문 데이터를 Excel로 다운로드 (옵션)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        # 현재 날짜 정보
        now = timezone.now()
        current_year = now.year
        current_month = now.month
        
        # 이번달 범위
        current_month_start = timezone.datetime(current_year, current_month, 1, tzinfo=timezone.get_current_timezone())
        if current_month == 12:
            current_month_end = timezone.datetime(current_year + 1, 1, 1, tzinfo=timezone.get_current_timezone())
        else:
            current_month_end = timezone.datetime(current_year, current_month + 1, 1, tzinfo=timezone.get_current_timezone())
        
        # 지난달 범위
        last_month_date = now - relativedelta(months=1)
        last_month_year = last_month_date.year
        last_month_month = last_month_date.month
        last_month_start = timezone.datetime(last_month_year, last_month_month, 1, tzinfo=timezone.get_current_timezone())
        if last_month_month == 12:
            last_month_end = timezone.datetime(last_month_year + 1, 1, 1, tzinfo=timezone.get_current_timezone())
        else:
            last_month_end = timezone.datetime(last_month_year, last_month_month + 1, 1, tzinfo=timezone.get_current_timezone())
        
        # URL 파라미터에서 필터 정보 가져오기
        filter_type = request.GET.get('filter', 'this_month')  # all, this_month, last_month, custom
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        page_number = request.GET.get('page', 1)
        
        # 기본 쿼리셋
        order_items = OrderItem.objects.filter(
            product=product
        ).select_related('order')
        
        # 날짜 필터링 적용 (화면과 동일한 로직)
        if filter_type == 'this_month':
            order_items = order_items.filter(
                order__created_at__gte=current_month_start,
                order__created_at__lt=current_month_end
            )
            filter_name = f"{current_month}월"
        elif filter_type == 'last_month':
            order_items = order_items.filter(
                order__created_at__gte=last_month_start,
                order__created_at__lt=last_month_end
            )
            filter_name = f"{last_month_month}월"
        elif filter_type == 'custom' and start_date and end_date:
            try:
                start_datetime = timezone.datetime.strptime(start_date, '%Y-%m-%d')
                start_datetime = start_datetime.replace(tzinfo=timezone.get_current_timezone())
                end_datetime = timezone.datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime.replace(tzinfo=timezone.get_current_timezone()) + timedelta(days=1)  # 종료일 포함
                order_items = order_items.filter(
                    order__created_at__gte=start_datetime,
                    order__created_at__lt=end_datetime
                )
                filter_name = f"{start_date}~{end_date}"
            except ValueError:
                # 날짜 형식이 잘못된 경우 전체 조회
                filter_name = "전체"
        else:
            # 전체 필터
            filter_name = "전체"
        
        # 정렬 적용 (최신순)
        order_items = order_items.order_by('-order__created_at')
        
        # 페이지네이션이 적용된 경우 해당 페이지의 데이터만 가져오기 (전체 필터일 때만)
        if filter_type == 'all' and page_number:
            paginator = Paginator(order_items, 10)
            try:
                page_obj = paginator.get_page(page_number)
                order_items = page_obj.object_list
                filter_name += f"_페이지{page_number}"
            except:
                # 페이지 번호가 잘못된 경우 첫 페이지
                page_obj = paginator.get_page(1)
                order_items = page_obj.object_list
                filter_name += "_페이지1"
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{product.title} 주문목록"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        headers = [
            '주문번호', '주문자', '수량', '단가(sats)', '옵션가격(sats)', '총액(sats)',
            '주문일시', '상태', '배송상태', '선택옵션', '연락처', '이메일',
            '우편번호', '주소', '상세주소', '요청사항'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row, item in enumerate(order_items, 2):
            order = item.order
            
            # 선택된 옵션을 문자열로 변환
            options_str = ''
            if item.selected_options:
                options_list = []
                for key, value in item.selected_options.items():
                    options_list.append(f"{key}: {value}")
                options_str = ', '.join(options_list)
            
            ws.cell(row=row, column=1, value=order.order_number)
            ws.cell(row=row, column=2, value=order.buyer_name)
            ws.cell(row=row, column=3, value=item.quantity)
            ws.cell(row=row, column=4, value=item.product_price)
            ws.cell(row=row, column=5, value=item.options_price)
            ws.cell(row=row, column=6, value=item.total_price)
            ws.cell(row=row, column=7, value=timezone.localtime(order.created_at).strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=8, value=order.get_status_display())
            ws.cell(row=row, column=9, value=order.get_delivery_status_display())
            ws.cell(row=row, column=10, value=options_str)
            # 우편번호 처리 - 0으로 시작하는 경우 문자열로 보존
            postal_code = order.shipping_postal_code or ''
            if postal_code and postal_code.startswith('0'):
                postal_code = f'="{postal_code}"'  # Excel에서 문자열로 인식하도록 처리
            
            ws.cell(row=row, column=11, value=order.buyer_phone or '')
            ws.cell(row=row, column=12, value=order.buyer_email or '')
            ws.cell(row=row, column=13, value=postal_code)
            ws.cell(row=row, column=14, value=order.shipping_address or '')
            ws.cell(row=row, column=15, value=order.shipping_detail_address or '')
            ws.cell(row=row, column=16, value=order.order_memo or '')
        
        # 열 너비 자동 조정
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # 메모리에 엑셀 파일 저장
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # HTTP 응답 생성
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 파일명에 필터 정보 포함
        filename = f"{store.store_id}_{product.title}_{filter_name}_주문목록.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        # openpyxl이 없으면 CSV로 fallback
        return export_product_orders_csv(request, store, product) 
